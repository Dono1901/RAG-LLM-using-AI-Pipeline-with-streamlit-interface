"""
Excel Processing Engine for RAG-LLM Financial System.
Handles xlsx, xlsm, xls, and csv files with intelligent sheet combination.
"""

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import pandas as pd
import numpy as np

from config import settings

logger = logging.getLogger(__name__)


@dataclass
class ColumnInfo:
    """Information about a detected column."""
    name: str
    dtype: str
    sample_values: List[Any]
    null_count: int
    unique_count: int
    is_numeric: bool
    is_date: bool
    is_currency: bool
    financial_category: Optional[str] = None


@dataclass
class SheetData:
    """Data from a single Excel sheet."""
    name: str
    df: pd.DataFrame
    header_row: int
    columns: List[ColumnInfo]
    detected_type: Optional[str] = None
    metadata: Dict[str, Any] = field(default_factory=dict)


@dataclass
class WorkbookData:
    """Data from an entire workbook."""
    file_path: Path
    sheets: List[SheetData]
    metadata: Dict[str, Any] = field(default_factory=dict)


@dataclass
class CombinedData:
    """Combined data from multiple sheets."""
    merged_df: Optional[pd.DataFrame]
    separate_sheets: List[SheetData]
    relationships: List[Dict[str, Any]]
    metadata: Dict[str, Any] = field(default_factory=dict)


@dataclass
class FinancialTable:
    """A detected financial table structure."""
    table_type: str  # income_statement, balance_sheet, cash_flow, budget, custom
    df: pd.DataFrame
    time_periods: List[str]
    line_items: List[str]
    source_sheet: str
    confidence: float


@dataclass
class DocumentChunk:
    """A chunk of document ready for RAG indexing."""
    text_content: str
    sheet_name: str
    table_structure: Optional[Dict[str, Any]]
    detected_financial_type: Optional[str]
    cell_references: Optional[str]
    metadata: Dict[str, Any] = field(default_factory=dict)


class ExcelProcessor:
    """
    Handles all Excel file operations with intelligent format detection.
    Supports: xlsx, xlsm, xls, csv, tsv
    """

    SUPPORTED_EXTENSIONS = {'.xlsx', '.xlsm', '.xls', '.csv', '.tsv'}

    # Financial column patterns for auto-detection
    REVENUE_PATTERNS = [
        r'revenue', r'sales', r'income', r'turnover', r'net\s*sales',
        r'gross\s*sales', r'total\s*revenue', r'top\s*line'
    ]
    EXPENSE_PATTERNS = [
        r'expense', r'cost', r'cogs', r'opex', r'sg&a', r'operating\s*expense',
        r'admin', r'selling', r'depreciation', r'amortization'
    ]
    ASSET_PATTERNS = [
        r'asset', r'cash', r'inventory', r'receivable', r'ppe',
        r'property', r'equipment', r'investment', r'prepaid'
    ]
    LIABILITY_PATTERNS = [
        r'liability', r'payable', r'debt', r'loan', r'accrued',
        r'deferred', r'obligation', r'mortgage'
    ]
    EQUITY_PATTERNS = [
        r'equity', r'capital', r'retained', r'shareholders',
        r'common\s*stock', r'treasury', r'accumulated'
    ]
    DATE_PATTERNS = [
        r'date', r'period', r'year', r'month', r'quarter', r'fy\d{2,4}',
        r'q[1-4]', r'jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec'
    ]

    def __init__(self, documents_path: str = "./documents"):
        """Initialize the Excel processor."""
        self.documents_path = Path(documents_path)
        self.loaded_workbooks: Dict[str, WorkbookData] = {}

    def scan_for_excel_files(self) -> List[Path]:
        """Scan documents folder for all Excel/CSV files."""
        files = []
        for ext in self.SUPPORTED_EXTENSIONS:
            files.extend(self.documents_path.glob(f"*{ext}"))
            files.extend(self.documents_path.glob(f"*{ext.upper()}"))
        return sorted(files, key=lambda x: x.stat().st_mtime, reverse=True)

    def load_workbook(self, file_path: Path) -> WorkbookData:
        """
        Load workbook with auto-detection of format.
        """
        file_path = Path(file_path)
        suffix = file_path.suffix.lower()

        logger.info(f"Loading workbook: {file_path.name}")

        sheets = []

        try:
            if suffix in ['.xlsx', '.xlsm']:
                sheets = self._load_openpyxl(file_path)
            elif suffix == '.xls':
                sheets = self._load_xlrd(file_path)
            elif suffix in ['.csv', '.tsv']:
                sheets = self._load_csv(file_path, delimiter=',' if suffix == '.csv' else '\t')
            else:
                raise ValueError(f"Unsupported file format: {suffix}")

            # Enforce row limits to prevent memory issues
            total_rows = sum(len(s.df) for s in sheets)
            if total_rows > settings.max_workbook_rows:
                logger.warning(
                    f"Workbook {file_path.name} has {total_rows} rows "
                    f"(limit {settings.max_workbook_rows}). Truncating sheets."
                )
                for sheet in sheets:
                    if len(sheet.df) > settings.max_workbook_rows:
                        sheet.df = sheet.df.head(settings.max_workbook_rows)

            # Detect financial types for each sheet
            for sheet in sheets:
                sheet.detected_type = self._detect_statement_type(sheet.df)
                sheet.columns = self._analyze_columns(sheet.df)

        except Exception as e:
            logger.error(f"Failed to load {file_path}: {e}")
            raise

        workbook = WorkbookData(
            file_path=file_path,
            sheets=sheets,
            metadata={
                'file_size': file_path.stat().st_size,
                'sheet_count': len(sheets),
                'total_rows': sum(len(s.df) for s in sheets)
            }
        )

        self.loaded_workbooks[str(file_path)] = workbook
        return workbook

    def _load_openpyxl(self, file_path: Path) -> List[SheetData]:
        """Load xlsx/xlsm files using openpyxl."""
        import openpyxl

        sheets = []
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data = list(ws.values)

            if not data:
                continue

            # Find header row
            header_row = self._find_header_row(data)

            # Create DataFrame
            if header_row >= 0 and header_row < len(data):
                headers = data[header_row]
                # Clean headers
                headers = [str(h).strip() if h else f"Column_{i}" for i, h in enumerate(headers)]
                df = pd.DataFrame(data[header_row + 1:], columns=headers)
            else:
                df = pd.DataFrame(data)

            # Clean the DataFrame
            df = self._clean_dataframe(df)

            if not df.empty:
                sheets.append(SheetData(
                    name=sheet_name,
                    df=df,
                    header_row=header_row,
                    columns=[],
                    metadata={'source_format': 'xlsx'}
                ))

        wb.close()
        return sheets

    def _load_xlrd(self, file_path: Path) -> List[SheetData]:
        """Load xls files using xlrd (or pandas as fallback)."""
        try:
            import xlrd
            wb = xlrd.open_workbook(file_path)
            sheets = []

            for sheet in wb.sheets():
                if sheet.nrows == 0:
                    continue

                data = [sheet.row_values(i) for i in range(sheet.nrows)]
                header_row = self._find_header_row(data)

                if header_row >= 0:
                    headers = [str(h).strip() if h else f"Column_{i}" for i, h in enumerate(data[header_row])]
                    df = pd.DataFrame(data[header_row + 1:], columns=headers)
                else:
                    df = pd.DataFrame(data)

                df = self._clean_dataframe(df)

                if not df.empty:
                    sheets.append(SheetData(
                        name=sheet.name,
                        df=df,
                        header_row=header_row,
                        columns=[],
                        metadata={'source_format': 'xls'}
                    ))

            return sheets
        except ImportError:
            # Fallback to pandas
            logger.warning("xlrd not available, using pandas for .xls files")
            return self._load_with_pandas(file_path)

    def _load_csv(self, file_path: Path, delimiter: str = ',') -> List[SheetData]:
        """Load CSV/TSV files."""
        try:
            # Try to detect encoding
            encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']

            df = None
            for encoding in encodings:
                try:
                    df = pd.read_csv(file_path, delimiter=delimiter, encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue

            if df is None:
                df = pd.read_csv(file_path, delimiter=delimiter, encoding='utf-8', errors='ignore')

            df = self._clean_dataframe(df)

            return [SheetData(
                name=file_path.stem,
                df=df,
                header_row=0,
                columns=[],
                metadata={'source_format': 'csv' if delimiter == ',' else 'tsv'}
            )]
        except Exception as e:
            logger.error(f"Failed to load CSV: {e}")
            return []

    def _load_with_pandas(self, file_path: Path) -> List[SheetData]:
        """Fallback loader using pandas."""
        sheets = []
        try:
            excel_file = pd.ExcelFile(file_path)
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                df = self._clean_dataframe(df)
                if not df.empty:
                    sheets.append(SheetData(
                        name=sheet_name,
                        df=df,
                        header_row=0,
                        columns=[],
                        metadata={'source_format': 'pandas'}
                    ))
        except Exception as e:
            logger.error(f"Pandas fallback failed: {e}")
        return sheets

    def _find_header_row(self, data: List[List]) -> int:
        """Find the most likely header row in the data."""
        if not data:
            return -1

        for i, row in enumerate(data[:10]):  # Check first 10 rows
            if not row:
                continue

            # Count non-empty cells
            non_empty = sum(1 for cell in row if cell is not None and str(cell).strip())

            # Check if row looks like a header (text values, no pure numbers)
            text_count = sum(
                1 for cell in row
                if cell is not None and isinstance(cell, str) and not cell.replace('.', '').replace('-', '').isdigit()
            )

            # If most cells are text and row is reasonably populated
            if non_empty >= 2 and text_count >= non_empty * 0.5:
                return i

        return 0  # Default to first row

    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and normalize a DataFrame."""
        if df.empty:
            return df

        # Remove completely empty rows and columns
        df = df.dropna(how='all', axis=0)
        df = df.dropna(how='all', axis=1)

        # Clean column names
        df.columns = [str(col).strip() for col in df.columns]

        # Convert numeric columns
        if len(df) == 0:
            return df

        for col in df.columns:
            try:
                numeric_col = pd.to_numeric(df[col], errors='coerce')
                if numeric_col.notna().sum() / len(df) > 0.5:
                    df[col] = numeric_col
            except Exception:
                pass

        return df

    def _analyze_columns(self, df: pd.DataFrame) -> List[ColumnInfo]:
        """Analyze columns and their characteristics."""
        columns = []

        for col in df.columns:
            series = df[col]

            # Determine if numeric
            is_numeric = pd.api.types.is_numeric_dtype(series)

            # Check for date
            is_date = False
            if series.dtype == 'object':
                date_pattern = re.compile(self.DATE_PATTERNS[0], re.IGNORECASE)
                if date_pattern.search(str(col)):
                    is_date = True

            # Check for currency (contains $ or numbers with commas)
            is_currency = False
            if series.dtype == 'object':
                sample = series.dropna().head(5).astype(str)
                is_currency = any('$' in str(v) or re.match(r'^[\d,]+\.?\d*$', str(v)) for v in sample)

            # Detect financial category
            financial_category = self._categorize_column(col)

            columns.append(ColumnInfo(
                name=col,
                dtype=str(series.dtype),
                sample_values=series.dropna().head(3).tolist(),
                null_count=series.isna().sum(),
                unique_count=series.nunique(),
                is_numeric=is_numeric,
                is_date=is_date,
                is_currency=is_currency,
                financial_category=financial_category
            ))

        return columns

    def _categorize_column(self, col_name: str) -> Optional[str]:
        """Categorize a column based on its name."""
        col_lower = col_name.lower()

        for pattern in self.REVENUE_PATTERNS:
            if re.search(pattern, col_lower, re.IGNORECASE):
                return 'revenue'

        for pattern in self.EXPENSE_PATTERNS:
            if re.search(pattern, col_lower, re.IGNORECASE):
                return 'expense'

        for pattern in self.ASSET_PATTERNS:
            if re.search(pattern, col_lower, re.IGNORECASE):
                return 'asset'

        for pattern in self.LIABILITY_PATTERNS:
            if re.search(pattern, col_lower, re.IGNORECASE):
                return 'liability'

        for pattern in self.EQUITY_PATTERNS:
            if re.search(pattern, col_lower, re.IGNORECASE):
                return 'equity'

        return None

    def _detect_statement_type(self, df: pd.DataFrame) -> Optional[str]:
        """Detect the type of financial statement."""
        if df.empty:
            return None

        # Combine all text for analysis
        text = ' '.join(df.columns.astype(str).tolist())
        if len(df) > 0:
            text += ' ' + ' '.join(df.iloc[:, 0].dropna().astype(str).tolist()[:20])

        text_lower = text.lower()

        # Score each statement type
        scores = {
            'income_statement': 0,
            'balance_sheet': 0,
            'cash_flow_statement': 0,
            'budget': 0
        }

        # Income statement indicators
        income_indicators = ['revenue', 'sales', 'gross profit', 'operating income',
                            'net income', 'ebitda', 'eps', 'earnings']
        scores['income_statement'] = sum(1 for ind in income_indicators if ind in text_lower)

        # Balance sheet indicators
        balance_indicators = ['assets', 'liabilities', 'equity', 'total assets',
                             'current assets', 'fixed assets', 'shareholders']
        scores['balance_sheet'] = sum(1 for ind in balance_indicators if ind in text_lower)

        # Cash flow indicators
        cf_indicators = ['cash flow', 'operating activities', 'investing activities',
                        'financing activities', 'free cash flow', 'capex']
        scores['cash_flow_statement'] = sum(1 for ind in cf_indicators if ind in text_lower)

        # Budget indicators
        budget_indicators = ['budget', 'forecast', 'actual', 'variance',
                            'plan', 'target', 'projected']
        scores['budget'] = sum(1 for ind in budget_indicators if ind in text_lower)

        # Return the highest scoring type (if score > 1)
        max_score = max(scores.values())
        if max_score > 1:
            return max(scores, key=scores.get)

        return 'custom'

    def combine_sheets_intelligently(self, workbook: WorkbookData) -> CombinedData:
        """
        Analyze all sheets and combine related data.
        """
        if not workbook.sheets:
            return CombinedData(
                merged_df=None,
                separate_sheets=[],
                relationships=[],
                metadata={}
            )

        # Group sheets by detected type
        sheets_by_type: Dict[str, List[SheetData]] = {}
        for sheet in workbook.sheets:
            sheet_type = sheet.detected_type or 'unknown'
            if sheet_type not in sheets_by_type:
                sheets_by_type[sheet_type] = []
            sheets_by_type[sheet_type].append(sheet)

        # Find relationships between sheets
        relationships = self._find_relationships(workbook.sheets)

        # Try to merge sheets with same schema
        merged_df = None
        mergeable_sheets = []

        for sheets in sheets_by_type.values():
            if len(sheets) > 1:
                # Check if schemas are similar enough to merge
                if self._can_merge_sheets(sheets):
                    merged = pd.concat([s.df for s in sheets], ignore_index=True)
                    merged['_source_sheet'] = [s.name for s in sheets for _ in range(len(s.df))]
                    mergeable_sheets.extend(sheets)
                    if merged_df is None:
                        merged_df = merged
                    else:
                        # Keep as separate if different types
                        pass

        # Keep non-merged sheets separate
        separate_sheets = [s for s in workbook.sheets if s not in mergeable_sheets]

        return CombinedData(
            merged_df=merged_df,
            separate_sheets=separate_sheets,
            relationships=relationships,
            metadata={
                'total_sheets': len(workbook.sheets),
                'merged_count': len(mergeable_sheets),
                'separate_count': len(separate_sheets),
                'detected_types': list(sheets_by_type.keys())
            }
        )

    def _can_merge_sheets(self, sheets: List[SheetData]) -> bool:
        """Check if sheets have similar enough schemas to merge."""
        if len(sheets) < 2:
            return False

        base_columns = set(sheets[0].df.columns)

        for sheet in sheets[1:]:
            sheet_columns = set(sheet.df.columns)
            max_cols = max(len(base_columns), len(sheet_columns))
            if max_cols == 0:
                return False
            overlap = len(base_columns & sheet_columns) / max_cols
            if overlap < 0.7:
                return False

        return True

    def _find_relationships(self, sheets: List[SheetData]) -> List[Dict[str, Any]]:
        """Find relationships between sheets based on common columns."""
        relationships = []

        for i, sheet1 in enumerate(sheets):
            for sheet2 in sheets[i + 1:]:
                common_cols = set(sheet1.df.columns) & set(sheet2.df.columns)
                if common_cols:
                    relationships.append({
                        'sheet1': sheet1.name,
                        'sheet2': sheet2.name,
                        'common_columns': list(common_cols),
                        'relationship_type': 'shared_columns'
                    })

        return relationships

    def extract_financial_tables(self, sheet_data: SheetData) -> List[FinancialTable]:
        """Identify and extract financial data structures."""
        tables = []

        df = sheet_data.df
        if df.empty:
            return tables

        # Extract time periods from columns
        time_periods = []
        for col in df.columns:
            if self._is_time_period(col):
                time_periods.append(col)

        # Extract line items from first column
        line_items = []
        if len(df.columns) > 0:
            first_col = df.iloc[:, 0].dropna().astype(str).tolist()
            line_items = [item for item in first_col if item.strip()]

        tables.append(FinancialTable(
            table_type=sheet_data.detected_type or 'custom',
            df=df,
            time_periods=time_periods,
            line_items=line_items[:50],  # Limit for memory
            source_sheet=sheet_data.name,
            confidence=0.8 if sheet_data.detected_type else 0.5
        ))

        return tables

    def _is_time_period(self, col_name: str) -> bool:
        """Check if a column name represents a time period."""
        col_lower = str(col_name).lower()

        time_patterns = [
            r'^\d{4}$',  # Year: 2024
            r'^q[1-4]\s*\d{4}$',  # Quarter: Q1 2024
            r'^fy\s*\d{2,4}$',  # Fiscal year: FY24
            r'^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)',  # Month names
            r'^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$',  # Date formats
        ]

        return any(re.match(pattern, col_lower) for pattern in time_patterns)

    def to_rag_chunks(self, data: CombinedData, workbook: WorkbookData,
                     chunk_size: int = 1000) -> List[DocumentChunk]:
        """
        Convert Excel data to RAG-compatible chunks.
        """
        chunks = []
        file_name = workbook.file_path.name

        # Process merged data
        if data.merged_df is not None and not data.merged_df.empty:
            merged_chunks = self._dataframe_to_chunks(
                data.merged_df,
                sheet_name="merged_data",
                file_name=file_name,
                chunk_size=chunk_size
            )
            chunks.extend(merged_chunks)

        # Process separate sheets
        for sheet in data.separate_sheets:
            sheet_chunks = self._dataframe_to_chunks(
                sheet.df,
                sheet_name=sheet.name,
                file_name=file_name,
                chunk_size=chunk_size,
                detected_type=sheet.detected_type
            )
            chunks.extend(sheet_chunks)

        # Add summary chunk for the workbook
        summary_chunk = self._create_summary_chunk(workbook, data)
        chunks.insert(0, summary_chunk)

        return chunks

    def _dataframe_to_chunks(self, df: pd.DataFrame, sheet_name: str,
                            file_name: str, chunk_size: int = 1000,
                            detected_type: Optional[str] = None) -> List[DocumentChunk]:
        """Convert a DataFrame to text chunks."""
        chunks = []

        if df.empty:
            return chunks

        # Create a text representation of the data
        # First, create a structured summary
        summary_lines = [
            f"Data from sheet: {sheet_name}",
            f"Columns ({len(df.columns)}): {', '.join(df.columns[:20])}",
            f"Rows: {len(df)}",
        ]

        if detected_type:
            summary_lines.append(f"Detected type: {detected_type}")

        summary_text = '\n'.join(summary_lines)

        # Convert data to markdown table format (limited rows per chunk)
        rows_per_chunk = max(10, chunk_size // 100)

        for start_idx in range(0, len(df), rows_per_chunk):
            end_idx = min(start_idx + rows_per_chunk, len(df))
            chunk_df = df.iloc[start_idx:end_idx]

            # Create markdown table
            try:
                table_text = chunk_df.to_markdown(index=False)
            except:
                # Fallback to string representation
                table_text = chunk_df.to_string(index=False)

            chunk_text = f"{summary_text}\n\nData (rows {start_idx + 1}-{end_idx}):\n{table_text}"

            chunks.append(DocumentChunk(
                text_content=chunk_text,
                sheet_name=sheet_name,
                table_structure={
                    'columns': list(df.columns),
                    'row_range': [start_idx, end_idx],
                    'total_rows': len(df)
                },
                detected_financial_type=detected_type,
                cell_references=f"{sheet_name}!A{start_idx + 2}:Z{end_idx + 1}",
                metadata={
                    'source_file': file_name,
                    'chunk_index': start_idx // rows_per_chunk
                }
            ))

        return chunks

    def _create_summary_chunk(self, workbook: WorkbookData,
                             combined: CombinedData) -> DocumentChunk:
        """Create a summary chunk for the entire workbook."""
        summary_lines = [
            f"Excel Workbook: {workbook.file_path.name}",
            f"Total sheets: {len(workbook.sheets)}",
            "",
            "Sheet Summary:"
        ]

        for sheet in workbook.sheets:
            summary_lines.append(
                f"  - {sheet.name}: {len(sheet.df)} rows, {len(sheet.df.columns)} columns, "
                f"type: {sheet.detected_type or 'unknown'}"
            )

        if combined.relationships:
            summary_lines.append("\nRelationships detected:")
            for rel in combined.relationships[:5]:
                summary_lines.append(
                    f"  - {rel['sheet1']} <-> {rel['sheet2']}: "
                    f"common columns: {', '.join(rel['common_columns'][:3])}"
                )

        # Add column summary
        all_columns = set()
        for sheet in workbook.sheets:
            all_columns.update(sheet.df.columns)

        summary_lines.append(f"\nAll columns across sheets: {', '.join(list(all_columns)[:30])}")

        return DocumentChunk(
            text_content='\n'.join(summary_lines),
            sheet_name="_summary",
            table_structure={'type': 'workbook_summary'},
            detected_financial_type=None,
            cell_references=None,
            metadata={
                'source_file': workbook.file_path.name,
                'is_summary': True
            }
        )

    def get_sheet_as_dataframe(self, file_path: Path, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """Get a specific sheet as a DataFrame."""
        workbook = self.loaded_workbooks.get(str(file_path))
        if not workbook:
            workbook = self.load_workbook(file_path)

        if sheet_name:
            for sheet in workbook.sheets:
                if sheet.name == sheet_name:
                    return sheet.df
            raise ValueError(f"Sheet '{sheet_name}' not found")

        # Return first sheet if no name specified
        if workbook.sheets:
            return workbook.sheets[0].df

        return pd.DataFrame()

    def get_all_numeric_columns(self, workbook: WorkbookData) -> Dict[str, List[str]]:
        """Get all numeric columns grouped by sheet."""
        result = {}
        for sheet in workbook.sheets:
            numeric_cols = [col.name for col in sheet.columns if col.is_numeric]
            if numeric_cols:
                result[sheet.name] = numeric_cols
        return result


def process_excel_for_rag(file_path: Path, documents_path: str = "./documents") -> List[Dict[str, Any]]:
    """
    Convenience function to process an Excel file and return RAG-compatible documents.
    """
    processor = ExcelProcessor(documents_path)
    workbook = processor.load_workbook(file_path)
    combined = processor.combine_sheets_intelligently(workbook)
    chunks = processor.to_rag_chunks(combined, workbook)

    return [
        {
            "content": chunk.text_content,
            "metadata": {
                "source": str(file_path.name),
                "sheet": chunk.sheet_name,
                "type": "excel",
                "table_structure": chunk.table_structure,
                "financial_type": chunk.detected_financial_type
            }
        }
        for chunk in chunks
    ]
